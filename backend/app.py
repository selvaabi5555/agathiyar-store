from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import openpyxl
import os
import io
import json
import requests

app = Flask(__name__, static_folder='uploads', static_url_path='/uploads')
CORS(app, supports_credentials=True)

# Config
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///agathiyar.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JWT_SECRET_KEY'] = 'agathiyar-store-secret-2024'
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=7)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB max

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

db = SQLAlchemy(app)
jwt = JWTManager(app)

# ─── MODELS ───────────────────────────────────────────────────────────────────

class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)

class Settings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    whatsapp_number = db.Column(db.String(20), default='')
    shop_name = db.Column(db.String(100), default='அகத்தியர் Store')
    shop_address = db.Column(db.String(200), default='C.Ayyampalayam, Manachnallur, Tamil Nadu 621005')
    shop_phone = db.Column(db.String(20), default='074483 67291')

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(50))
    price = db.Column(db.Float, nullable=False)
    stock = db.Column(db.Integer, default=0)
    image_url = db.Column(db.String(300), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Staff(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20))
    role = db.Column(db.String(50))
    joined_at = db.Column(db.DateTime, default=datetime.utcnow)

class Bill(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bill_number = db.Column(db.String(20), unique=True, nullable=False)
    customer_name = db.Column(db.String(100), nullable=False)
    customer_phone = db.Column(db.String(20))
    items = db.Column(db.Text, nullable=False)  # JSON string
    subtotal = db.Column(db.Float, nullable=False)
    discount = db.Column(db.Float, default=0)
    total = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(20), nullable=False)  # cash/gpay/credit
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'), nullable=True)

class Review(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_name = db.Column(db.String(100), nullable=False)
    rating = db.Column(db.Integer, nullable=False)
    comment = db.Column(db.Text)
    approved = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ─── INIT DB ──────────────────────────────────────────────────────────────────

def init_db():
    with app.app_context():
        db.create_all()
        # Default admin
        if not Admin.query.first():
            db.session.add(Admin(username='admin', password='admin123'))
        # Default settings
        if not Settings.query.first():
            db.session.add(Settings())
        # Sample reviews
        if not Review.query.first():
            db.session.add_all([
                Review(customer_name='Mani Kandan', rating=5, comment='நல்லா dress கிடைக்கும் 🦋🙏', approved=True),
                Review(customer_name='P SELVA', rating=5, comment='அருமையான கடை 🎯🙌', approved=True),
            ])
        db.session.commit()

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def generate_bill_number():
    count = Bill.query.count() + 1
    return f"AGT-{datetime.now().year}-{count:04d}"

def get_settings():
    return Settings.query.first()

def send_whatsapp_pdf(bill, pdf_bytes):
    settings = get_settings()
    number = settings.whatsapp_number if settings else ''
    if not number:
        return False
    try:
        # WhatsApp Web link (opens in browser for manual send)
        msg = f"🏪 {settings.shop_name}\n🧾 Bill No: {bill.bill_number}\n👤 Customer: {bill.customer_name}\n💰 Total: ₹{bill.total}\n💳 Payment: {bill.payment_method.upper()}\n📅 {bill.created_at.strftime('%d-%m-%Y %H:%M')}"
        return True
    except:
        return False

# ─── AUTH ROUTES ──────────────────────────────────────────────────────────────

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    admin = Admin.query.filter_by(username=data.get('username')).first()
    if admin and admin.password == data.get('password'):
        token = create_access_token(identity=str(admin.id))
        return jsonify({'token': token, 'username': admin.username})
    return jsonify({'error': 'தவறான username அல்லது password'}), 401

@app.route('/api/verify', methods=['GET'])
@jwt_required()
def verify():
    return jsonify({'valid': True})

# ─── SETTINGS ROUTES ──────────────────────────────────────────────────────────

@app.route('/api/settings', methods=['GET'])
def get_settings_route():
    s = get_settings()
    return jsonify({
        'shop_name': s.shop_name,
        'shop_address': s.shop_address,
        'shop_phone': s.shop_phone,
        'whatsapp_number': s.whatsapp_number
    })

@app.route('/api/settings', methods=['PUT'])
@jwt_required()
def update_settings():
    data = request.json
    s = get_settings()
    s.shop_name = data.get('shop_name', s.shop_name)
    s.shop_address = data.get('shop_address', s.shop_address)
    s.shop_phone = data.get('shop_phone', s.shop_phone)
    s.whatsapp_number = data.get('whatsapp_number', s.whatsapp_number)
    db.session.commit()
    return jsonify({'success': True})

# ─── PRODUCT ROUTES ───────────────────────────────────────────────────────────

@app.route('/api/products', methods=['GET'])
def get_products():
    products = Product.query.all()
    return jsonify([{
        'id': p.id, 'name': p.name, 'category': p.category,
        'price': p.price, 'stock': p.stock, 'image_url': p.image_url or ''
    } for p in products])

@app.route('/api/products', methods=['POST'])
@jwt_required()
def add_product():
    data = request.json
    p = Product(name=data['name'], category=data.get('category', ''),
                price=data['price'], stock=data.get('stock', 0),
                image_url=data.get('image_url', ''))
    db.session.add(p)
    db.session.commit()
    return jsonify({'success': True, 'id': p.id})

@app.route('/api/products/<int:pid>', methods=['PUT'])
@jwt_required()
def update_product(pid):
    p = Product.query.get_or_404(pid)
    data = request.json
    p.name = data.get('name', p.name)
    p.category = data.get('category', p.category)
    p.price = data.get('price', p.price)
    p.stock = data.get('stock', p.stock)
    p.image_url = data.get('image_url', p.image_url)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/products/<int:pid>', methods=['DELETE'])
@jwt_required()
def delete_product(pid):
    p = Product.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    return jsonify({'success': True})

# ─── STAFF ROUTES ─────────────────────────────────────────────────────────────

@app.route('/api/staff', methods=['GET'])
@jwt_required()
def get_staff():
    staff = Staff.query.all()
    return jsonify([{
        'id': s.id, 'name': s.name, 'phone': s.phone,
        'role': s.role, 'joined_at': s.joined_at.strftime('%d-%m-%Y')
    } for s in staff])

@app.route('/api/staff', methods=['POST'])
@jwt_required()
def add_staff():
    data = request.json
    s = Staff(name=data['name'], phone=data.get('phone', ''), role=data.get('role', ''))
    db.session.add(s)
    db.session.commit()
    return jsonify({'success': True, 'id': s.id})

@app.route('/api/staff/<int:sid>', methods=['DELETE'])
@jwt_required()
def delete_staff(sid):
    s = Staff.query.get_or_404(sid)
    db.session.delete(s)
    db.session.commit()
    return jsonify({'success': True})

# ─── BILL ROUTES ──────────────────────────────────────────────────────────────

@app.route('/api/bills', methods=['GET'])
@jwt_required()
def get_bills():
    bills = Bill.query.order_by(Bill.created_at.desc()).all()
    return jsonify([{
        'id': b.id, 'bill_number': b.bill_number, 'customer_name': b.customer_name,
        'customer_phone': b.customer_phone, 'total': b.total,
        'payment_method': b.payment_method, 'created_at': b.created_at.strftime('%d-%m-%Y %H:%M'),
        'discount': b.discount
    } for b in bills])

@app.route('/api/bills', methods=['POST'])
@jwt_required()
def create_bill():
    data = request.json
    items = data['items']
    subtotal = sum(item['price'] * item['qty'] for item in items)
    discount = data.get('discount', 0)
    total = subtotal - discount

    # Update stock
    for item in items:
        p = Product.query.get(item['product_id'])
        if p and p.stock >= item['qty']:
            p.stock -= item['qty']

    bill = Bill(
        bill_number=generate_bill_number(),
        customer_name=data['customer_name'],
        customer_phone=data.get('customer_phone', ''),
        items=json.dumps(items),
        subtotal=subtotal,
        discount=discount,
        total=total,
        payment_method=data['payment_method'],
        staff_id=data.get('staff_id')
    )
    db.session.add(bill)
    db.session.commit()

    # Generate PDF and get WhatsApp link
    pdf_bytes = generate_bill_pdf(bill)
    wa_link = get_whatsapp_link(bill)

    return jsonify({
        'success': True,
        'bill_number': bill.bill_number,
        'bill_id': bill.id,
        'whatsapp_link': wa_link
    })

@app.route('/api/bills/<int:bid>/pdf', methods=['GET'])
def get_bill_pdf(bid):
    bill = Bill.query.get_or_404(bid)
    pdf_bytes = generate_bill_pdf(bill)
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'Bill-{bill.bill_number}.pdf'
    )

def get_whatsapp_link(bill):
    settings = get_settings()
    number = (settings.whatsapp_number or '').replace('+', '').replace(' ', '')
    if not number:
        return None
    msg = f"🏪 {settings.shop_name}%0A🧾 Bill: {bill.bill_number}%0A👤 {bill.customer_name}%0A💰 Total: ₹{bill.total}%0A💳 {bill.payment_method.upper()}%0A📅 {bill.created_at.strftime('%d-%m-%Y')}"
    return f"https://wa.me/{number}?text={msg}"

def generate_bill_pdf(bill):
    from reportlab.platypus import HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.lib.pagesizes import A5

    buffer = io.BytesIO()
    settings = get_settings()

    # Tamil -> English fallback for PDF (ReportLab doesn't support Tamil)
    shop_name_raw = settings.shop_name or 'Agathiyar Store'
    # Remove non-ASCII (Tamil) chars, keep English
    import re
    shop_name_en = re.sub(r'[^-]+', '', shop_name_raw).strip()
    if not shop_name_en:
        shop_name_en = 'Agathiyar Store'
    shop_name_upper = shop_name_en.upper()

    doc = SimpleDocTemplate(
        buffer, pagesize=A5,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1*cm, bottomMargin=1*cm
    )
    story = []

    MAROON    = colors.HexColor('#7A1A2E')
    GOLD      = colors.HexColor('#C9972A')
    CREAM     = colors.HexColor('#FBF6EE')
    LIGHTCREAM= colors.HexColor('#F5EDE0')
    DARKTEXT  = colors.HexColor('#1A0A0E')
    MUTEDTEXT = colors.HexColor('#7A6050')
    WHITE     = colors.white
    A5W = A5[0] - 3*cm

    # ── HEADER ──────────────────────────────────────────────
    def P(text, **kw):
        return Paragraph(text, ParagraphStyle('x', **kw))

    header = Table([
        [P(f"<b>{shop_name_upper}</b>",
           fontName='Helvetica-Bold', fontSize=20, textColor=WHITE,
           alignment=1, leading=24)],
        [P(settings.shop_address or 'C.Ayyampalayam, Manachanallur, Tamil Nadu 621005',
           fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#F5DEB3'),
           alignment=1, leading=12)],
        [P(f"Ph: {settings.shop_phone or '074483 67291'}  |  Open Daily till 9:30 PM",
           fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#F5DEB3'),
           alignment=1, leading=12)],
    ], colWidths=[A5W])
    header.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), MAROON),
        ('TOPPADDING',(0,0),(-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
        ('LEFTPADDING',(0,0),(-1,-1), 12),
        ('RIGHTPADDING',(0,0),(-1,-1), 12),
    ]))
    story.append(header)
    story.append(HRFlowable(width='100%', thickness=3, color=GOLD, spaceBefore=0, spaceAfter=10))

    # ── INVOICE TITLE ────────────────────────────────────────
    story.append(P('<b>INVOICE</b>',
        fontName='Helvetica-Bold', fontSize=13, textColor=MAROON, leading=16, spaceAfter=8))

    # ── BILL INFO ────────────────────────────────────────────
    lbl = dict(fontName='Helvetica-Bold', fontSize=9, textColor=DARKTEXT, leading=13)
    val = dict(fontName='Helvetica', fontSize=9, textColor=DARKTEXT, leading=13)
    date_str = bill.created_at.strftime('%d/%m/%Y, %I:%M:%S %p').lower()

    info = Table([
        [P('<b>Bill No</b>',**lbl), P(f' :  {bill.bill_number}',**val),
         P('<b>Customer</b>',**lbl), P(f' :  {bill.customer_name}',**val)],
        [P('<b>Date</b>',**lbl),    P(f' :  {date_str}',**val),
         P('<b>Phone</b>',**lbl),   P(f' :  {bill.customer_phone or "—"}',**val)],
    ], colWidths=[2*cm, 5*cm, 2.4*cm, 4*cm])
    info.setStyle(TableStyle([
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(info)
    story.append(Spacer(1, 0.35*cm))

    # ── ITEMS TABLE ───────────────────────────────────────────
    th = dict(fontName='Helvetica-Bold', fontSize=9, textColor=WHITE, leading=12)
    td = dict(fontName='Helvetica', fontSize=9, textColor=DARKTEXT, leading=12)
    tdr = dict(fontName='Helvetica', fontSize=9, textColor=DARKTEXT, leading=12, alignment=2)
    total_lbl = dict(fontName='Helvetica-Bold', fontSize=9, textColor=DARKTEXT, leading=12, alignment=2)
    total_val = dict(fontName='Helvetica-Bold', fontSize=10, textColor=MAROON, leading=13, alignment=2)
    disc_val  = dict(fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#C0392B'), leading=12, alignment=2)

    items = json.loads(bill.items)
    n = len(items)

    rows = [[P('Item',**th), P('Qty',**th), P('Unit Price',**th), P('Amount',**th)]]
    for i, item in enumerate(items):
        amt = item['price'] * item['qty']
        rows.append([
            P(item['name'], **td),
            P(str(item['qty']), **td),
            P(f"Rs.{item['price']:,.0f}", **tdr),
            P(f"Rs.{amt:,.0f}", **tdr),
        ])

    # Subtotal / Discount / Total
    if bill.discount > 0:
        rows.append(['','', P('Subtotal', **total_lbl), P(f"Rs.{bill.subtotal:,.0f}", **tdr)])
        rows.append(['','', P('Discount', **total_lbl), P(f"-Rs.{bill.discount:,.0f}", **disc_val)])
    rows.append(['','', P('<b>TOTAL AMOUNT:</b>', **total_lbl), P(f"<b>Rs.{bill.total:,.0f}</b>", **total_val)])

    col_w = [7*cm, 1.2*cm, 2.6*cm, 2.6*cm]
    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    row_bg = []
    for r in range(1, n+1):
        row_bg.append(('BACKGROUND',(0,r),(-1,r), WHITE if r%2==1 else LIGHTCREAM))

    tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0), MAROON),
        ('TOPPADDING',(0,0),(-1,-1),5),
        ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('LEFTPADDING',(0,0),(-1,-1),6),
        ('RIGHTPADDING',(0,0),(-1,-1),6),
        ('BOX',(0,0),(-1,n),0.5,colors.HexColor('#D4C0A8')),
        ('LINEBELOW',(0,0),(-1,0),0.5,GOLD),
        ('LINEABOVE',(2,n+1),(-1,n+1),0.8,colors.HexColor('#D4C0A8')),
        ('BACKGROUND',(0,n+1),(-1,-1), CREAM),
        ('TOPPADDING',(0,n+1),(-1,-1),4),
        ('BOTTOMPADDING',(0,n+1),(-1,-1),4),
    ] + row_bg))
    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── FOOTER ───────────────────────────────────────────────
    footer = Table([
        [P(f"<i>Thank you for shopping at {shop_name_en} Store!</i>",
           fontName='Helvetica-BoldOblique', fontSize=9,
           textColor=MAROON, alignment=1, leading=13)],
        [P(f"For enquiries: {settings.shop_phone or '074483 67291'}  |  {settings.shop_address or 'C.Ayyampalayam, Manachanallur'}",
           fontName='Helvetica', fontSize=7,
           textColor=MUTEDTEXT, alignment=1, leading=11)],
    ], colWidths=[A5W])
    footer.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), CREAM),
        ('TOPPADDING',(0,0),(-1,-1),8),
        ('BOTTOMPADDING',(0,0),(-1,-1),8),
        ('LEFTPADDING',(0,0),(-1,-1),10),
        ('RIGHTPADDING',(0,0),(-1,-1),10),
        ('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#D4C0A8')),
    ]))
    story.append(footer)

    doc.build(story)
    return buffer.getvalue()



@app.route('/api/bills/<int:bid>', methods=['DELETE'])
@jwt_required()
def delete_bill(bid):
    b = Bill.query.get_or_404(bid)
    db.session.delete(b)
    db.session.commit()
    return jsonify({'success': True})

# ─── REVIEW ROUTES ────────────────────────────────────────────────────────────

@app.route('/api/reviews', methods=['GET'])
def get_reviews():
    reviews = Review.query.filter_by(approved=True).order_by(Review.created_at.desc()).all()
    return jsonify([{
        'id': r.id, 'customer_name': r.customer_name, 'rating': r.rating,
        'comment': r.comment, 'created_at': r.created_at.strftime('%d-%m-%Y')
    } for r in reviews])

@app.route('/api/reviews/all', methods=['GET'])
@jwt_required()
def get_all_reviews():
    reviews = Review.query.order_by(Review.created_at.desc()).all()
    return jsonify([{
        'id': r.id, 'customer_name': r.customer_name, 'rating': r.rating,
        'comment': r.comment, 'approved': r.approved,
        'created_at': r.created_at.strftime('%d-%m-%Y')
    } for r in reviews])

@app.route('/api/reviews', methods=['POST'])
def add_review():
    data = request.json
    r = Review(customer_name=data['customer_name'], rating=data['rating'],
                comment=data.get('comment', ''))
    db.session.add(r)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/reviews/<int:rid>/approve', methods=['PUT'])
@jwt_required()
def approve_review(rid):
    r = Review.query.get_or_404(rid)
    r.approved = not r.approved
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/reviews/<int:rid>', methods=['DELETE'])
@jwt_required()
def delete_review(rid):
    r = Review.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'success': True})

# ─── REPORTS ──────────────────────────────────────────────────────────────────

@app.route('/api/reports/excel', methods=['GET'])
@jwt_required()
def export_excel():
    bills = Bill.query.order_by(Bill.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ['Bill No', 'Customer', 'Phone', 'Items', 'Subtotal', 'Discount', 'Total', 'Payment', 'Date']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='8B1A1A')
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')

    for b in bills:
        items = json.loads(b.items)
        items_str = ', '.join(f"{i['name']}x{i['qty']}" for i in items)
        ws.append([b.bill_number, b.customer_name, b.customer_phone, items_str,
                   b.subtotal, b.discount, b.total, b.payment_method,
                   b.created_at.strftime('%d-%m-%Y %H:%M')])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Sales-Report.xlsx')

@app.route('/api/dashboard', methods=['GET'])
@jwt_required()
def dashboard():
    today = datetime.now().date()
    today_bills = Bill.query.filter(db.func.date(Bill.created_at) == today).all()
    total_bills = Bill.query.count()
    total_revenue = db.session.query(db.func.sum(Bill.total)).scalar() or 0
    today_revenue = sum(b.total for b in today_bills)
    low_stock = Product.query.filter(Product.stock < 5).count()

    return jsonify({
        'total_bills': total_bills,
        'total_revenue': round(total_revenue, 2),
        'today_revenue': round(today_revenue, 2),
        'today_bills': len(today_bills),
        'low_stock': low_stock,
        'total_products': Product.query.count(),
        'total_staff': Staff.query.count(),
        'pending_reviews': Review.query.filter_by(approved=False).count()
    })


# ─── DATABASE VIEWER ──────────────────────────────────────────────────────────

@app.route('/api/db/tables', methods=['GET'])
@jwt_required()
def db_tables():
    return jsonify(['products', 'staff', 'bills', 'reviews', 'settings', 'admin'])

@app.route('/api/db/<table>', methods=['GET'])
@jwt_required()
def db_view(table):
    try:
        if table == 'products':
            rows = Product.query.all()
            data = [{'id':r.id,'name':r.name,'category':r.category,'price':r.price,'stock':r.stock,'image_url':r.image_url or ''} for r in rows]
        elif table == 'staff':
            rows = Staff.query.all()
            data = [{'id':r.id,'name':r.name,'phone':r.phone,'role':r.role,'joined_at':str(r.joined_at)} for r in rows]
        elif table == 'bills':
            rows = Bill.query.order_by(Bill.created_at.desc()).all()
            data = [{'id':r.id,'bill_number':r.bill_number,'customer_name':r.customer_name,'customer_phone':r.customer_phone,'subtotal':r.subtotal,'discount':r.discount,'total':r.total,'payment_method':r.payment_method,'created_at':str(r.created_at)} for r in rows]
        elif table == 'reviews':
            rows = Review.query.all()
            data = [{'id':r.id,'customer_name':r.customer_name,'rating':r.rating,'comment':r.comment,'approved':r.approved,'created_at':str(r.created_at)} for r in rows]
        elif table == 'settings':
            rows = Settings.query.all()
            data = [{'id':r.id,'shop_name':r.shop_name,'shop_address':r.shop_address,'shop_phone':r.shop_phone,'whatsapp_number':r.whatsapp_number} for r in rows]
        elif table == 'admin':
            rows = Admin.query.all()
            data = [{'id':r.id,'username':r.username,'password':'***'} for r in rows]
        else:
            return jsonify({'error':'Table not found'}), 404
        return jsonify({'table': table, 'count': len(data), 'rows': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)

# ─── IMAGE UPLOAD ─────────────────────────────────────────────────────────────

@app.route('/api/upload', methods=['POST'])
@jwt_required()
def upload_image():
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type'}), 400
    import uuid
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
    return jsonify({'url': f'http://localhost:5000/uploads/{filename}'})
